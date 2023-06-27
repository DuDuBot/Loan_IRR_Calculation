import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def IRR_calculation_clear_formula(file_path, sheet_IRR_calculation):
    """
        Clear old data in IRR sheet
    """
    book = load_workbook(file_path)
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    sheet = writer.sheets[sheet_IRR_calculation]
    start_row = 2
    end_row = sheet.max_row
    start_col = 4
    end_col = sheet.max_column

    for row in sheet[start_row: end_row + 1]:
        for cell in row[start_col: end_col + 1]:
            cell.value = None
            cell.border = None
            cell.fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
    writer.save()


def ppmt(rate, period, total_period, pv, fv=0, pmt_type=0):
    """
    Calculates the principal payment for a specific period in a loan or investment.
    """
    payment = -pv * rate / (1 - (1 + rate) ** -total_period)
    interest_pmt = -pv * rate
    principal_pmt = payment - interest_pmt
    if pmt_type == 1:
        principal_pmt /= (1 + rate)
    for _ in range(1, period):
        pv += principal_pmt
        interest_pmt = -pv * rate
        principal_pmt = payment - interest_pmt
        if pmt_type == 1:
            principal_pmt /= (1 + rate)
    return principal_pmt


def pmt(rate, total_period, pv, fv=0, pmt_type=0):
    """
    Calculates the fixed periodic payment for a loan or investment.
    """
    if rate == 0:
        payment = (fv - pv) / total_period
    else:
        payment = rate * (fv + pv * (1 + rate) ** total_period) / ((1 + rate) ** total_period - 1)
    if pmt_type == 1:
        payment /= (1 + rate)
    return -payment


def calculate_irr(cashflows):
    """
    Calculates the Internal Rate of Return (IRR) for a series of cashflows.
    """
    def npv(rate):
        return sum(c / (1 + rate) ** i for i, c in enumerate(cashflows))

    def npv_derivative(rate):
        return sum(-i * c / (1 + rate) ** (i + 1) for i, c in enumerate(cashflows))

    # Newton-Raphson iteration
    epsilon = 1e-6
    max_iterations = 100
    irr = 0.1  # initial guess is 0.1

    for _ in range(max_iterations):
        npv_value = npv(irr)
        npv_derivative_value = npv_derivative(irr)
        new_irr = irr - npv_value / npv_derivative_value

        if abs(new_irr - irr) < epsilon:
            return new_irr

        irr = new_irr

    # if max iteration exceed
    raise ValueError("IRR did not converge, please adjust max iterations or tolerance.")


def write_IRR_results(file_path, sheet_IRR_calculation, df, irr_annualized):
    """
        Write result dataframe and IRR to original sheet
    """
    start_row = 1  # first row
    start_col = 4  # E column
    book = load_workbook(file_path)
    writer = pd.ExcelWriter(file_path, engine='openpyxl', mode='a')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    sheet = writer.sheets[sheet_IRR_calculation]

    for row in range(df.shape[0]):
        for col in range(df.shape[1]):
            cell = sheet.cell(row=row + start_row + 1, column=col + start_col + 1)
            cell.value = df.iloc[row, col]

    IRR_cell = sheet.cell(row=df.shape[0] + start_row, column=df.shape[1] + start_col + 2)
    IRR_result_cell = IRR_cell.offset(0, 1)
    IRR_cell.value = 'IRR ='
    IRR_result_cell.value = irr_annualized

    writer.save()




